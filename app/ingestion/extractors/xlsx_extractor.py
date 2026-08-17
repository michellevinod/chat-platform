from pathlib import Path

from openpyxl import load_workbook

from app.ingestion.extractors.base_extractor import BaseExtractor
from app.ingestion.extractors.raw_models import (
    RawBlock,
    RawDocument,
    RawPage,
    RawTableBlock,
    RawTextBlock,
)
from app.models.enums import BlockType


class XLSXExtractor(BaseExtractor):
    """
    Extracts text and tables from Microsoft Excel (.xlsx) workbooks.

    Each worksheet is treated as one page.
    """

    def extract(
        self,
        file_path: Path,
    ) -> RawDocument:

        workbook = load_workbook(
            filename=file_path,
            data_only=True,
        )

        pages: list[RawPage] = []

        page_number = 1

        for sheet in workbook.worksheets:

            blocks: list[RawBlock] = []

            block_number = 0

            # Sheet title
            blocks.append(
                RawTextBlock(
                    page_number=page_number,
                    block_number=block_number,
                    block_type=BlockType.TEXT,
                    bbox=(0.0, 0.0, 0.0, 0.0),
                    text=f"Worksheet: {sheet.title}",
                )
            )

            block_number += 1

            rows_data = []
            for row in sheet.iter_rows(values_only=True):

                values = []

                for cell in row:
                    if cell is None:
                        values.append("")
                    else:
                        values.append(str(cell).strip())

                if any(v for v in values):
                    rows_data.append(values)

                    blocks.append(
                        RawTextBlock(
                            page_number=page_number,
                            block_number=block_number,
                            block_type=BlockType.TEXT,
                            bbox=(0.0, 0.0, 0.0, 0.0),
                            text=" | ".join([v for v in values if v]),
                        )
                    )

                    block_number += 1

            if rows_data:
                headers = rows_data[0]
                lines = []
                lines.append("| " + " | ".join(headers) + " |")
                lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
                for r in rows_data[1:]:
                    lines.append("| " + " | ".join(r) + " |")
                md_table = "\n".join(lines)

                blocks.append(
                    RawTableBlock(
                        page_number=page_number,
                        block_number=block_number,
                        block_type=BlockType.TABLE,
                        bbox=(0.0, 0.0, 0.0, 0.0),
                        markdown=f"### Worksheet: {sheet.title}\n\n{md_table}",
                        rows=rows_data,
                        headers=headers,
                        caption=f"Worksheet: {sheet.title}",
                    )
                )
                block_number += 1

            pages.append(
                RawPage(
                    page_number=page_number,
                    blocks=blocks,
                )
            )

            page_number += 1

        return RawDocument(
            file_name=file_path.name,
            total_pages=len(pages),
            pages=pages,
        )